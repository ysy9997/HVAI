import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
# from openpyxl.drawing import Image as OpenpyxlImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image
from pathlib import Path
import os
from typing import List, Dict, Tuple
import io

class ExcelImagePredictionGenerator:
    def __init__(self, csv_files: List[str], model_names: List[str], 
                 image_folder: str, class_names: List[str] = None):
        """
        Initialize the Excel generator.
        
        Args:
            csv_files: List of paths to CSV files with predictions
            model_names: List of model names corresponding to CSV files
            image_folder: Path to folder containing images
            class_names: List of class names (if None, will read from CSV header)
        """
        self.csv_files = csv_files
        self.model_names = model_names
        self.image_folder = image_folder
        self.class_names = class_names
        self.predictions = {}
        self.image_names = None
        
    def load_predictions(self):
        """Load prediction data from CSV files."""
        print("Loading prediction files...")
        
        for i, (file_path, model_name) in enumerate(zip(self.csv_files, self.model_names)):
            df = pd.read_csv(file_path)
            
            # If class names not provided, read from first row (header)
            if self.class_names is None and i == 0:
                # First column is image names, rest are class names
                self.class_names = df.columns[1:].tolist()
                print(f"Detected {len(self.class_names)} classes: {self.class_names[:5]}...")
            
            # First column is image names
            if i == 0:
                self.image_names = df.iloc[:, 0].values
            
            # Extract probability predictions (columns 1 onwards)
            prob_matrix = df.iloc[:, 1:].values
            self.predictions[model_name] = prob_matrix
            
            print(f"Loaded {model_name}: {prob_matrix.shape[0]} samples, {prob_matrix.shape[1]} classes")
    
    def get_top1_predictions(self) -> Dict[str, List[Tuple[str, float]]]:
        """Get top-1 predictions with confidence for each model."""
        print("Extracting top-1 predictions...")
        
        top1_predictions = {}
        
        for model_name, preds in self.predictions.items():
            model_top1 = []
            
            for i in range(len(preds)):
                # Get the class with highest probability
                top_class_idx = np.argmax(preds[i])
                top_confidence = preds[i][top_class_idx]
                top_class_name = self.class_names[top_class_idx]
                
                model_top1.append((top_class_name, top_confidence))
            
            top1_predictions[model_name] = model_top1
            print(f"{model_name}: Extracted {len(model_top1)} top-1 predictions")
        
        return top1_predictions
    
    def resize_image(self, image_path: str, max_size: Tuple[int, int] = (100, 100)) -> str:
        """
        Resize image to fit in Excel cell and return path to resized image.
        
        Args:
            image_path: Path to original image
            max_size: Maximum size (width, height) for the resized image
            
        Returns:
            Path to resized image
        """
        resized_dir = str(Path(image_path).parent) + '_resized'
        os.makedirs(resized_dir, exist_ok=True)
        
        try:
            temp_path = image_path.replace('.jpg', '_resized.jpg').replace('.png', '_resized.png')
            temp_path = resized_dir + '/' + Path(image_path).name
            if not os.path.exists(temp_path):
                with Image.open(image_path) as img:
                    # Convert to RGB if necessary
                    if img.mode != 'RGB':
                        img = img.convert('RGB')
                    
                    # Calculate resize ratio maintaining aspect ratio
                    img_width, img_height = img.size
                    max_width, max_height = max_size
                    
                    ratio = min(max_width / img_width, max_height / img_height)
                    new_width = int(img_width * ratio)
                    new_height = int(img_height * ratio)
                    
                    # Resize image
                    resized_img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                    
                    # Save resized image to temporary location
                    resized_img.save(temp_path, quality=85)
                
            return temp_path
                
        except Exception as e:
            print(f"Error resizing image {image_path}: {e}")
            return None
    
    def generate_excel(self, output_path: str, image_size: Tuple[int, int] = (80, 80),
                      row_height: int = 60, img_col_width: int = 12):
        """
        Generate Excel file with embedded images and predictions.
        
        Args:
            output_path: Path for output Excel file
            image_size: Size for embedded images (width, height)
            row_height: Height of each row in pixels
            img_col_width: Width of image column
        """
        print(f"Generating Excel file: {output_path}")
        
        # Load predictions
        self.load_predictions()
        top1_predictions = self.get_top1_predictions()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Model Predictions Comparison"
        
        # Define headers - add all_agree as leftmost column
        headers = ['all_agree', 'Image Name', 'Image'] + [model for model in self.model_names]
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type='solid')
        
        # Set column widths
        ws.column_dimensions['A'].width = 10  # all_agree column
        ws.column_dimensions['B'].width = 15  # Image name column
        ws.column_dimensions['C'].width = img_col_width  # Image column
        for i in range(len(self.model_names)):
            col_letter = get_column_letter(4 + i)
            ws.column_dimensions[col_letter].width = 20  # Prediction columns
        
        # Set header row height
        ws.row_dimensions[1].height = 30
        
        # Process each image and add to Excel
        missing_images = []
        processed_count = 0
        
        for idx, img_name in enumerate(self.image_names):
            row_num = idx + 2  # +2 because Excel is 1-indexed and we have a header
            
            # Set row height for image display
            ws.row_dimensions[row_num].height = row_height
            
            # Check if all models predict the same class
            predicted_classes = [top1_predictions[model][idx][0] for model in self.model_names]
            all_agree = 1 if len(set(predicted_classes)) == 1 else 0
            
            # Column A: all_agree indicator
            cell = ws.cell(row=row_num, column=1)
            cell.value = all_agree
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # Color code: green if all agree, red if disagree
            if all_agree == 1:
                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
            else:
                cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Light pink
            
            # Column B: Image name
            ws.cell(row=row_num, column=2).value = img_name
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center', vertical='center')
            
            # Column C: Embedded image
            img_path = os.path.join(self.image_folder, img_name + ".jpg")
            
            if os.path.exists(img_path):
                try:
                    # Resize image
                    resized_img_path = self.resize_image(img_path, image_size)
                    
                    if resized_img_path:
                        # Create image object and add to worksheet
                        img_obj = openpyxl.drawing.image.Image(resized_img_path)
                        
                        # Position image in cell
                        cell_address = f'C{row_num}'
                        img_obj.anchor = cell_address
                        
                        # Add image to worksheet
                        ws.add_image(img_obj)
                        
                        # # Clean up temporary resized image
                        # if resized_img_path != img_path:
                        #     try:
                        #         os.remove(resized_img_path)
                        #     except:
                        #         pass
                                
                except Exception as e:
                    print(f"Error adding image {img_name}: {e}")
                    ws.cell(row=row_num, column=3).value = "Image Error"
                    missing_images.append(img_name)
            else:
                ws.cell(row=row_num, column=3).value = "Image Not Found"
                missing_images.append(img_name)
            
            # Columns D onwards: Model predictions
            for col_idx, model_name in enumerate(self.model_names, 4):
                class_name, confidence = top1_predictions[model_name][idx]
                prediction_text = f"{class_name}: {confidence:.3f}"
                
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = prediction_text
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Color coding based on confidence
                if confidence >= 0.9:
                    cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
                elif confidence >= 0.7:
                    cell.fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')  # Light yellow
                else:
                    cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Light pink
            
            processed_count += 1
            if processed_count % 50 == 0:
                print(f"Processed {processed_count}/{len(self.image_names)} samples...")
        
        # Add summary sheet
        self.add_summary_sheet(wb, top1_predictions, missing_images)
        
        # Save workbook
        try:
            wb.save(output_path)
            print(f"\nExcel file saved successfully: {output_path}")
            print(f"Processed {processed_count} samples")
            if missing_images:
                print(f"Warning: {len(missing_images)} images were not found")
        except Exception as e:
            print(f"Error saving Excel file: {e}")
    
    def add_summary_sheet(self, workbook: Workbook, top1_predictions: Dict, missing_images: List):
        """Add a summary sheet with statistics."""
        ws_summary = workbook.create_sheet("Summary")
        
        # Title
        ws_summary.cell(row=1, column=1).value = "Model Predictions Summary"
        ws_summary.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        row = 3
        
        # Dataset info
        ws_summary.cell(row=row, column=1).value = "Dataset Information:"
        ws_summary.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        ws_summary.cell(row=row, column=1).value = f"Total samples: {len(self.image_names)}"
        row += 1
        ws_summary.cell(row=row, column=1).value = f"Number of classes: {len(self.class_names)}"
        row += 1
        ws_summary.cell(row=row, column=1).value = f"Number of models: {len(self.model_names)}"
        row += 1
        ws_summary.cell(row=row, column=1).value = f"Missing images: {len(missing_images)}"
        row += 2
        
        # Model statistics
        ws_summary.cell(row=row, column=1).value = "Model Statistics:"
        ws_summary.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        # Headers for model stats
        headers = ["Model", "Avg Confidence", "High Conf (>0.9)", "Medium Conf (0.7-0.9)", "Low Conf (<0.7)"]
        for col, header in enumerate(headers, 1):
            ws_summary.cell(row=row, column=col).value = header
            ws_summary.cell(row=row, column=col).font = Font(bold=True)
        row += 1
        
        # Calculate statistics for each model
        for model_name in self.model_names:
            predictions = top1_predictions[model_name]
            confidences = [conf for _, conf in predictions]
            
            avg_conf = np.mean(confidences)
            high_conf = sum(1 for conf in confidences if conf >= 0.9)
            med_conf = sum(1 for conf in confidences if 0.7 <= conf < 0.9)
            low_conf = sum(1 for conf in confidences if conf < 0.7)
            
            ws_summary.cell(row=row, column=1).value = model_name
            ws_summary.cell(row=row, column=2).value = f"{avg_conf:.3f}"
            ws_summary.cell(row=row, column=3).value = f"{high_conf} ({high_conf/len(predictions)*100:.1f}%)"
            ws_summary.cell(row=row, column=4).value = f"{med_conf} ({med_conf/len(predictions)*100:.1f}%)"
            ws_summary.cell(row=row, column=5).value = f"{low_conf} ({low_conf/len(predictions)*100:.1f}%)"
            row += 1
        
        # Set column widths for summary
        for col in range(1, 6):
            ws_summary.column_dimensions[get_column_letter(col)].width = 15

def generate_prediction_excel(csv_files: List[str], model_names: List[str], 
                            image_folder: str, output_path: str,
                            class_names: List[str] = None,
                            image_size: Tuple[int, int] = (80, 80)):
    """
    Main function to generate Excel file with images and predictions.
    
    Args:
        csv_files: List of paths to CSV files with predictions
        model_names: List of model names
        image_folder: Path to folder containing images
        output_path: Path for output Excel file
        class_names: Optional list of class names
        image_size: Size for embedded images (width, height)
    """
    
    generator = ExcelImagePredictionGenerator(
        csv_files=csv_files,
        model_names=model_names,
        image_folder=image_folder,
        class_names=class_names
    )
    
    generator.generate_excel(
        output_path=output_path,
        image_size=image_size,
        row_height=65,  # Adjust based on image size
        img_col_width=12
    )
    
    return generator

# Example usage
if __name__ == "__main__":
    # Example configuration
    BASE = r"C:\Users\Acer\vscode"
    
    csv_files = [
        f'{BASE}/hvai_submissions/250528_baseline_0.3517.csv',
        f'{BASE}/hvai_submissions/250531_convnextbase_0.2501763867.csv',
        f'{BASE}/hvai_submissions/250531_v11x_0.2864.csv',
        f'{BASE}/hvai_submissions/geometric_mean_ensemble_predictions.csv',
    ]
    model_names = ['ResNet18', 'ConvNext', 'YOLOv11', 'Ensemble']
    
    image_folder = f'{BASE}/open/test'  # Folder containing TEST_0000.jpg, TEST_0001.jpg, etc.
    
    output_path = 'model_predictions_with_images.xlsx'
    
    # Generate Excel file
    print("Starting Excel generation...")
    generator = generate_prediction_excel(
        csv_files=csv_files,
        model_names=model_names,
        image_folder=image_folder,
        output_path=output_path,
        class_names=None,  # Will read from CSV header
        image_size=(240, 240)  # Adjust image size as needed
    )
    
    print("Excel generation completed!")
    
    # Optional: Print some statistics
    print(f"\nGenerated Excel with:")
    print(f"- {len(generator.image_names)} samples")
    print(f"- {len(generator.class_names)} classes") 
    print(f"- {len(generator.model_names)} models")